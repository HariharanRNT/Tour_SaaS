"""
File content extraction service for the itinerary import flow.

Handles: PDF (PyMuPDF → pdfplumber → Gemini Vision OCR), DOCX (python-docx),
         XLSX/XLS (openpyxl), TXT (UTF-8).

NOTE: Legacy binary .doc files are NOT supported — python-docx only reads the
      modern .docx (OOXML) format. Users uploading .doc files receive a clear
      message asking them to re-save as .docx in Word/LibreOffice.

This module is intentionally standalone so it can be tested independently.
"""
from __future__ import annotations

import io
import logging
from typing import Tuple

logger = logging.getLogger(__name__)

# ─── Constants ────────────────────────────────────────────────────────────────

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

ACCEPTED_EXTENSIONS = {".pdf", ".docx", ".xls", ".xlsx", ".txt"}

# If extracted text is below this character count after two PDF passes, treat
# the file as a scanned/image PDF and trigger OCR via Gemini Vision.
PDF_OCR_THRESHOLD = 100


# ─── Validation ───────────────────────────────────────────────────────────────

def validate_file(filename: str, size: int) -> None:
    """Raise ValueError with a user-friendly message on bad input."""
    ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if size > MAX_FILE_SIZE:
        mb = size / 1024 / 1024
        raise ValueError(f"File size exceeds the 10 MB limit ({mb:.1f} MB uploaded).")
    if ext == ".doc":
        raise ValueError(
            "Legacy .doc files are not supported. Please open the file in Microsoft Word or "
            "LibreOffice Writer and save it as .docx, then upload again."
        )
    if ext not in ACCEPTED_EXTENSIONS:
        raise ValueError(
            "Only PDF, Word (.docx), Excel (.xls/.xlsx), and Text (.txt) files are supported."
        )


# ─── PDF extraction ───────────────────────────────────────────────────────────

def _extract_pdf_pymupdf(data: bytes) -> str:
    """Primary PDF extraction via PyMuPDF (fitz)."""
    import fitz  # PyMuPDF

    text_parts: list[str] = []
    with fitz.open(stream=data, filetype="pdf") as doc:
        for page in doc:
            text_parts.append(page.get_text("text"))
    return "\n\n".join(text_parts)


def _extract_pdf_pdfplumber(data: bytes) -> str:
    """Fallback PDF extraction via pdfplumber."""
    import pdfplumber

    text_parts: list[str] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            text_parts.append(page_text)
    return "\n\n".join(text_parts)


async def _extract_pdf_ocr_gemini(data: bytes) -> str:
    """
    OCR fallback: send each PDF page as an image to Gemini Vision.
    Used only when both PyMuPDF and pdfplumber return < PDF_OCR_THRESHOLD chars.
    """
    import fitz  # PyMuPDF (used to rasterise pages)
    from google import genai
    from google.genai import types as gtypes
    import base64
    from app.config import settings

    logger.info("[FileExtractor] Scanned PDF detected — using Gemini Vision OCR")

    api_key = getattr(settings, "GEMINI_API_KEY", None) or getattr(
        settings, "GOOGLE_AI_API_KEY", None
    )
    if not api_key:
        raise RuntimeError("No Gemini API key configured for OCR fallback.")

    client = genai.Client(api_key=api_key)

    page_texts: list[str] = []
    with fitz.open(stream=data, filetype="pdf") as doc:
        for page_num, page in enumerate(doc, start=1):
            # Render page to PNG at 150 DPI
            mat = fitz.Matrix(150 / 72, 150 / 72)
            pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
            png_bytes = pix.tobytes("png")
            b64 = base64.b64encode(png_bytes).decode()

            response = client.models.generate_content(
                model="gemini-2.0-flash",
                contents=[
                    gtypes.Part.from_bytes(
                        data=base64.b64decode(b64),
                        mime_type="image/png",
                    ),
                    "Extract ALL text visible in this image exactly as written. Output only the raw text, no commentary.",
                ],
            )
            page_text = response.text.strip() if response.text else ""
            logger.debug(f"[FileExtractor] OCR page {page_num}: {len(page_text)} chars")
            page_texts.append(page_text)

    return "\n\n".join(page_texts)


async def extract_pdf(data: bytes) -> str:
    """
    Extract text from a PDF.
    Order: PyMuPDF → pdfplumber → Gemini Vision OCR.
    """
    text = ""

    # 1. PyMuPDF
    try:
        text = _extract_pdf_pymupdf(data)
        logger.debug(f"[FileExtractor] PyMuPDF extracted {len(text)} chars")
    except Exception as e:
        logger.warning(f"[FileExtractor] PyMuPDF failed: {e}")

    # 2. pdfplumber fallback
    if len(text.strip()) < PDF_OCR_THRESHOLD:
        try:
            text = _extract_pdf_pdfplumber(data)
            logger.debug(f"[FileExtractor] pdfplumber extracted {len(text)} chars")
        except Exception as e:
            logger.warning(f"[FileExtractor] pdfplumber failed: {e}")

    # 3. Gemini Vision OCR for scanned PDFs
    if len(text.strip()) < PDF_OCR_THRESHOLD:
        try:
            text = await _extract_pdf_ocr_gemini(data)
            logger.debug(f"[FileExtractor] Gemini OCR extracted {len(text)} chars")
        except Exception as e:
            logger.error(f"[FileExtractor] Gemini OCR failed: {e}")
            raise ValueError(
                "Unable to read this PDF. It appears to be a scanned document and OCR processing failed."
            )

    if not text.strip():
        raise ValueError("Unable to extract content from the uploaded PDF.")
    return text


# ─── DOCX extraction ──────────────────────────────────────────────────────────

def extract_docx(data: bytes) -> str:
    """Extract text from a DOCX file using python-docx."""
    from docx import Document
    from docx.oxml.ns import qn

    doc = Document(io.BytesIO(data))
    parts: list[str] = []

    for element in doc.element.body:
        # Headings and paragraphs
        if element.tag.endswith("}p"):
            from docx.text.paragraph import Paragraph
            para = Paragraph(element, doc)
            text = para.text.strip()
            if text:
                # Prefix headings so Gemini understands structure
                if para.style.name.startswith("Heading"):
                    parts.append(f"\n## {text}")
                else:
                    parts.append(text)
        # Tables
        elif element.tag.endswith("}tbl"):
            from docx.table import Table
            tbl = Table(element, doc)
            for row in tbl.rows:
                row_text = " | ".join(c.text.strip() for c in row.cells if c.text.strip())
                if row_text:
                    parts.append(row_text)

    result = "\n".join(parts)
    if not result.strip():
        raise ValueError("Unable to extract content from the uploaded Word document.")
    return result


# ─── Excel extraction ─────────────────────────────────────────────────────────

def extract_xlsx(data: bytes) -> str:
    """Extract text from an XLSX/XLS file using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet_parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                rows_text.append(" | ".join(cells))
        if rows_text:
            sheet_parts.append(f"Sheet: {sheet_name}\n" + "\n".join(rows_text))

    wb.close()
    result = "\n\n---\n\n".join(sheet_parts)
    if not result.strip():
        raise ValueError("Unable to extract content from the uploaded Excel file.")
    return result


# ─── TXT extraction ───────────────────────────────────────────────────────────

def extract_txt(data: bytes) -> str:
    """Read a plain text file as UTF-8 (with latin-1 fallback)."""
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError:
        return data.decode("latin-1", errors="replace")


# ─── Main entry point ─────────────────────────────────────────────────────────

async def extract_file_content(filename: str, data: bytes) -> str:
    """
    Validate the file and extract its text content.

    Args:
        filename: Original file name (used to detect extension).
        data:     Raw file bytes.

    Returns:
        Extracted plain-text string ready for Gemini.

    Raises:
        ValueError: User-friendly error message on any failure.
    """
    validate_file(filename, len(data))

    ext = "." + filename.rsplit(".", 1)[-1].lower()

    try:
        if ext == ".pdf":
            return await extract_pdf(data)
        elif ext == ".docx":
            return extract_docx(data)
        elif ext == ".doc":
            # Should never reach here — validate_file() rejects .doc before this point.
            # Guard included as a defence-in-depth measure.
            raise ValueError(
                "Legacy .doc files are not supported. Please re-save as .docx "
                "(File → Save As → Word Document) and upload again."
            )
        elif ext in (".xlsx", ".xls"):
            return extract_xlsx(data)
        elif ext == ".txt":
            return extract_txt(data)
        else:
            raise ValueError(
                "Only PDF, Word (.docx), Excel (.xls/.xlsx), and Text (.txt) files are supported."
            )
    except ValueError:
        raise  # pass user-friendly errors straight through
    except Exception as e:
        logger.error(f"[FileExtractor] Unexpected error extracting {filename}: {e}", exc_info=True)
        raise ValueError(f"Unable to extract content from the uploaded file: {str(e)}")
